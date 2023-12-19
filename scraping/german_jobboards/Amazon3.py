# sprawdzić employment type, może uda się znaleźć
# być może stripować (Aktualizacja) przy update,
# co zrobić z update? Nie pasuje do kategorii
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from bs4 import BeautifulSoup
import requests
import regex as re
import pandas as pd
from flashtext import KeywordProcessor
from unidecode import unidecode
from jobboard_scraper3 import JobboardScraper
from openpyxl import Workbook
from datetime import datetime, timedelta, date
# from r'path\jobboard_scraper_excel.py' import JobboardScraper
# import sys
# sys.path.append('.')
# from jobboard_scraper import JobboardScraper
path = '/path/to/your/chromedriver'
driver = webdriver.Chrome(path)


wb = Workbook()
ws = wb.active
class AmazonScraper(JobboardScraper):
    global ws
    
    def __init__(self, country, jobboard, currency):
        # self.session = requests.Session()
        # self.headers = {'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_10_1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/39.0.2171.95 Safari/537.36'}
        url = 'https://www.amazon.jobs/de/search?offset=0&result_limit=10&sort=relevant&cities[]=Munich%2C%20Bavaria%2C%20DEU&cities[]=Berlin%2C%20Berlin%2C%20DEU&cities[]=DEU&cities[]=Leipzig%2C%20Saxony%2C%20DEU&cities[]=Dresden%2C%20Saxony%2C%20DEU&cities[]=Schkeuditz%2C%20Saxony%2C%20DEU&cities[]=Frankfurt%2C%20Hesse%2C%20DEU&cities[]=Wissen%2C%20Rhineland-Palatinate%2C%20DEU&cities[]=Winsen%2C%20Schleswig-Holstein%2C%20DEU&cities[]=Wyhl%20am%20Kaiserstuhl%2C%20Baden-Wurttemberg%2C%20DEU&cities[]=Pforzheim%2C%20Baden-Wurttemberg%2C%20DEU&cities[]=S%C3%BClzetal%2C%20Sachsen-Anhalt%2C%20DEU&cities[]=Cologne%2C%20North-Rhine-Westphalia%2C%20DEU&cities[]=Dortmund%2C%20North-Rhine-Westphalia%2C%20DEU&cities[]=Frankenthal%20(Pfalz)%2C%20Rhineland-Palatinate%2C%20DEU&cities[]=Bad%20Hersfeld%2C%20Hesse%2C%20DEU&cities[]=Oelde%2C%20North-Rhine-Westphalia%2C%20DEU&cities[]=Winsen%20(Luhe)%2C%20Lower-Saxony%2C%20DEU&cities[]=Rheinberg%2C%20North-Rhine-Westphalia%2C%20DEU&cities[]=Graben%2C%20Bavaria%2C%20DEU&cities[]=Hamburg%2C%20Hamburg%2C%20DEU&cities[]=M%C3%B6nchengladbach%2C%20North-Rhine-Westphalia%2C%20DEU&cities[]=Wolfhagen%2C%20Hesse%2C%20DEU&cities[]=Bremen%2C%20Bremen%2C%20DEU&cities[]=Kobern-Gondorf%2C%20Rhineland-Palatinate%2C%20DEU&cities[]=Wunstorf%2C%20Lower-Saxony%2C%20DEU&cities[]=Werne%2C%20North-Rhine-Westphalia%2C%20DEU&cities[]=Krefeld%2C%20North-Rhine-Westphalia%2C%20DEU&cities[]=Sch%C3%B6nefeld%2C%20Brandenburg%2C%20DEU&cities[]=Aachen%2C%20North-Rhine-Westphalia%2C%20DEU&cities[]=Cottbus%2C%20Brandenburg%2C%20DEU&cities[]=Dusseldorf%2C%20North-Rhine-Westphalia%2C%20DEU&cities[]=Emsb%C3%BCren%2C%20Lower-Saxony%2C%20DEU&cities[]=Erfurt%2C%20Thuringia%2C%20DEU&cities[]=Gernsheim%2C%20Hesse%2C%20DEU&cities[]=Bielefeld%2C%20North-Rhine-Westphalia%2C%20DEU&cities[]=Cloppenburg%2C%20Lower-Saxony%2C%20DEU&cities[]=Eggolsheim%2C%20Bavaria%2C%20DEU&cities[]=Freiburg%2C%20Baden-Wurttemberg%2C%20DEU&cities[]=Neuwied%2C%20Rhineland-Palatinate%2C%20DEU&cities[]=Paderborn%2C%20North-Rhine-Westphalia%2C%20DEU&cities[]=T%C3%BCbingen%2C%20Baden-Wurttemberg%2C%20DEU&cities[]=Brieselang%2C%20Brandenburg%2C%20DEU&cities[]=Garbsen%2C%20Lower-Saxony%2C%20DEU&cities[]=Hanover%2C%20Lower-Saxony%2C%20DEU&cities[]=Raunheim%2C%20Hesse%2C%20DEU&cities[]=V%C3%B6lklingen%2C%20Saarland%2C%20DEU&cities[]=Witten%2C%20North-Rhine-Westphalia%2C%20DEU&cities[]=Emden%2C%20Lower-Saxony%2C%20DEU&cities[]=Mannheim%2C%20Baden-Wurttemberg%2C%20DEU&cities[]=Nuremberg%2C%20Bavaria%2C%20DEU&cities[]=Bad%20Oldesloe%2C%20Schleswig-Holstein%2C%20DEU&cities[]=Bochum%2C%20North-Rhine-Westphalia%2C%20DEU&cities[]=Borgstedt%2C%20Schleswig-Holstein%2C%20DEU&cities[]=Essen%2C%20North-Rhine-Westphalia%2C%20DEU&cities[]=Magdeburg%2C%20Sachsen-Anhalt%2C%20DEU&cities[]=Me%C3%9Fkirch%2C%20Baden-Wurttemberg%2C%20DEU&cities[]=Schortens%2C%20Lower-Saxony%2C%20DEU&cities[]=Trierweiler%2C%20Rhineland-Palatinate%2C%20DEU&cities[]=Bayreuth%2C%20Bavaria%2C%20DEU&cities[]=Duisburg%2C%20North-Rhine-Westphalia%2C%20DEU&cities[]=Echzell%2C%20Hesse%2C%20DEU&cities[]=Kn%C3%BCllwald%2C%20Hesse%2C%20DEU&cities[]=Lengede%2C%20Lower-Saxony%2C%20DEU&cities[]=Rostock%2C%20Mecklenburg-Vorpommern%2C%20DEU&cities[]=Unna%2C%20North-Rhine-Westphalia%2C%20DEU&cities[]=Frankenberg%20(Eder)%2C%20Hesse%2C%20DEU&cities[]=Grolsheim%2C%20Rhineland-Palatinate%2C%20DEU&cities[]=N%C3%BCtzen%2C%20Schleswig-Holstein%2C%20DEU&cities[]=Pommersfelden%2C%20Bavaria%2C%20DEU&cities[]=Sonneberg%2C%20Thuringia%2C%20DEU&cities[]=Waldorf%2C%20Rhineland-Palatinate%2C%20DEU&cities[]=Wuppertal%2C%20North-Rhine-Westphalia%2C%20DEU&cities[]=Eschweiler%2C%20North-Rhine-Westphalia%2C%20DEU&cities[]=Gersthofen%2C%20Bavaria%2C%20DEU&cities[]=Hoppegarten%2C%20Brandenburg%2C%20DEU&cities[]=Kempen%2C%20North-Rhine-Westphalia%2C%20DEU&cities[]=Stuttgart%2C%20Baden-Wurttemberg%2C%20DEU&cities[]=Ulm%2C%20Baden-Wurttemberg%2C%20DEU&distanceType=Mi&radius=24km&latitude=&longitude=&loc_group_id=&loc_query=&base_query=&city=&country=&region=&county=&query_options=&'
        driver.get(url)
        super().__init__(country, jobboard, currency)  

    def explore_ads(self):
        timeout = 10
        wait = WebDriverWait(driver, timeout)
        childframe = wait.until(EC.presence_of_element_located((By.TAG_NAME, "body")))

        sel_obj = driver.find_element_by_tag_name('body')
        ads_container_sel = sel_obj.find_element_by_xpath("//div[@class='search-content']//div[@class='container']")
        ads = ads_container_sel.find_element_by_class_name('pagination-control')
        all_ads_sel = ads.find_elements_by_class_name('page-button')
        button = all_ads_sel[-1].text
        pages = int(button)
        
        driver.close()
        nums = 2        # do numeracji wierszy
          
        ws['A1'] = 'Scraping Date'
        ws['B1'] = 'Company'
        ws['C1'] = 'Title'
        ws['D1'] = 'Company Name'
        ws['E1'] = 'Location'
        ws['F1'] = 'Date'
        ws['G1'] = 'Sector'
        ws['H1'] = 'Salary'
        ws['I1'] = 'Agency or Direct'
        ws['J1'] = 'Employment Type'
        ws['K1'] = 'Description'
        ws['L1'] = 'Skills'
        ws['M1'] = 'Link'
        ws['N1'] = 'Date of Update'
        # for page in range(0, pages):
        for page in range(0, 3):    # to jest tylko do testów, normalnie można dać range(0, pages)
            if page == 0:
                url = 'https://www.amazon.jobs/de/search?offset=0&result_limit=10&sort=relevant&cities[]=Munich%2C%20Bavaria%2C%20DEU&cities[]=Berlin%2C%20Berlin%2C%20DEU&cities[]=DEU&cities[]=Leipzig%2C%20Saxony%2C%20DEU&cities[]=Dresden%2C%20Saxony%2C%20DEU&cities[]=Schkeuditz%2C%20Saxony%2C%20DEU&cities[]=Frankfurt%2C%20Hesse%2C%20DEU&cities[]=Wissen%2C%20Rhineland-Palatinate%2C%20DEU&cities[]=Winsen%2C%20Schleswig-Holstein%2C%20DEU&cities[]=Wyhl%20am%20Kaiserstuhl%2C%20Baden-Wurttemberg%2C%20DEU&cities[]=Pforzheim%2C%20Baden-Wurttemberg%2C%20DEU&cities[]=S%C3%BClzetal%2C%20Sachsen-Anhalt%2C%20DEU&cities[]=Cologne%2C%20North-Rhine-Westphalia%2C%20DEU&cities[]=Dortmund%2C%20North-Rhine-Westphalia%2C%20DEU&cities[]=Frankenthal%20(Pfalz)%2C%20Rhineland-Palatinate%2C%20DEU&cities[]=Bad%20Hersfeld%2C%20Hesse%2C%20DEU&cities[]=Oelde%2C%20North-Rhine-Westphalia%2C%20DEU&cities[]=Winsen%20(Luhe)%2C%20Lower-Saxony%2C%20DEU&cities[]=Rheinberg%2C%20North-Rhine-Westphalia%2C%20DEU&cities[]=Graben%2C%20Bavaria%2C%20DEU&cities[]=Hamburg%2C%20Hamburg%2C%20DEU&cities[]=M%C3%B6nchengladbach%2C%20North-Rhine-Westphalia%2C%20DEU&cities[]=Wolfhagen%2C%20Hesse%2C%20DEU&cities[]=Bremen%2C%20Bremen%2C%20DEU&cities[]=Kobern-Gondorf%2C%20Rhineland-Palatinate%2C%20DEU&cities[]=Wunstorf%2C%20Lower-Saxony%2C%20DEU&cities[]=Werne%2C%20North-Rhine-Westphalia%2C%20DEU&cities[]=Krefeld%2C%20North-Rhine-Westphalia%2C%20DEU&cities[]=Sch%C3%B6nefeld%2C%20Brandenburg%2C%20DEU&cities[]=Aachen%2C%20North-Rhine-Westphalia%2C%20DEU&cities[]=Cottbus%2C%20Brandenburg%2C%20DEU&cities[]=Dusseldorf%2C%20North-Rhine-Westphalia%2C%20DEU&cities[]=Emsb%C3%BCren%2C%20Lower-Saxony%2C%20DEU&cities[]=Erfurt%2C%20Thuringia%2C%20DEU&cities[]=Gernsheim%2C%20Hesse%2C%20DEU&cities[]=Bielefeld%2C%20North-Rhine-Westphalia%2C%20DEU&cities[]=Cloppenburg%2C%20Lower-Saxony%2C%20DEU&cities[]=Eggolsheim%2C%20Bavaria%2C%20DEU&cities[]=Freiburg%2C%20Baden-Wurttemberg%2C%20DEU&cities[]=Neuwied%2C%20Rhineland-Palatinate%2C%20DEU&cities[]=Paderborn%2C%20North-Rhine-Westphalia%2C%20DEU&cities[]=T%C3%BCbingen%2C%20Baden-Wurttemberg%2C%20DEU&cities[]=Brieselang%2C%20Brandenburg%2C%20DEU&cities[]=Garbsen%2C%20Lower-Saxony%2C%20DEU&cities[]=Hanover%2C%20Lower-Saxony%2C%20DEU&cities[]=Raunheim%2C%20Hesse%2C%20DEU&cities[]=V%C3%B6lklingen%2C%20Saarland%2C%20DEU&cities[]=Witten%2C%20North-Rhine-Westphalia%2C%20DEU&cities[]=Emden%2C%20Lower-Saxony%2C%20DEU&cities[]=Mannheim%2C%20Baden-Wurttemberg%2C%20DEU&cities[]=Nuremberg%2C%20Bavaria%2C%20DEU&cities[]=Bad%20Oldesloe%2C%20Schleswig-Holstein%2C%20DEU&cities[]=Bochum%2C%20North-Rhine-Westphalia%2C%20DEU&cities[]=Borgstedt%2C%20Schleswig-Holstein%2C%20DEU&cities[]=Essen%2C%20North-Rhine-Westphalia%2C%20DEU&cities[]=Magdeburg%2C%20Sachsen-Anhalt%2C%20DEU&cities[]=Me%C3%9Fkirch%2C%20Baden-Wurttemberg%2C%20DEU&cities[]=Schortens%2C%20Lower-Saxony%2C%20DEU&cities[]=Trierweiler%2C%20Rhineland-Palatinate%2C%20DEU&cities[]=Bayreuth%2C%20Bavaria%2C%20DEU&cities[]=Duisburg%2C%20North-Rhine-Westphalia%2C%20DEU&cities[]=Echzell%2C%20Hesse%2C%20DEU&cities[]=Kn%C3%BCllwald%2C%20Hesse%2C%20DEU&cities[]=Lengede%2C%20Lower-Saxony%2C%20DEU&cities[]=Rostock%2C%20Mecklenburg-Vorpommern%2C%20DEU&cities[]=Unna%2C%20North-Rhine-Westphalia%2C%20DEU&cities[]=Frankenberg%20(Eder)%2C%20Hesse%2C%20DEU&cities[]=Grolsheim%2C%20Rhineland-Palatinate%2C%20DEU&cities[]=N%C3%BCtzen%2C%20Schleswig-Holstein%2C%20DEU&cities[]=Pommersfelden%2C%20Bavaria%2C%20DEU&cities[]=Sonneberg%2C%20Thuringia%2C%20DEU&cities[]=Waldorf%2C%20Rhineland-Palatinate%2C%20DEU&cities[]=Wuppertal%2C%20North-Rhine-Westphalia%2C%20DEU&cities[]=Eschweiler%2C%20North-Rhine-Westphalia%2C%20DEU&cities[]=Gersthofen%2C%20Bavaria%2C%20DEU&cities[]=Hoppegarten%2C%20Brandenburg%2C%20DEU&cities[]=Kempen%2C%20North-Rhine-Westphalia%2C%20DEU&cities[]=Stuttgart%2C%20Baden-Wurttemberg%2C%20DEU&cities[]=Ulm%2C%20Baden-Wurttemberg%2C%20DEU&distanceType=Mi&radius=24km&latitude=&longitude=&loc_group_id=&loc_query=&base_query=&city=&country=&region=&county=&query_options=&'
            else:
                url = 'https://www.amazon.jobs/de/search?offset=' + str(page) + '0&result_limit=10&sort=relevant&cities[]=Munich%2C%20Bavaria%2C%20DEU&cities[]=Berlin%2C%20Berlin%2C%20DEU&cities[]=DEU&cities[]=Leipzig%2C%20Saxony%2C%20DEU&cities[]=Dresden%2C%20Saxony%2C%20DEU&cities[]=Schkeuditz%2C%20Saxony%2C%20DEU&cities[]=Frankfurt%2C%20Hesse%2C%20DEU&cities[]=Wissen%2C%20Rhineland-Palatinate%2C%20DEU&cities[]=Winsen%2C%20Schleswig-Holstein%2C%20DEU&cities[]=Wyhl%20am%20Kaiserstuhl%2C%20Baden-Wurttemberg%2C%20DEU&cities[]=Pforzheim%2C%20Baden-Wurttemberg%2C%20DEU&cities[]=S%C3%BClzetal%2C%20Sachsen-Anhalt%2C%20DEU&cities[]=Cologne%2C%20North-Rhine-Westphalia%2C%20DEU&cities[]=Dortmund%2C%20North-Rhine-Westphalia%2C%20DEU&cities[]=Frankenthal%20(Pfalz)%2C%20Rhineland-Palatinate%2C%20DEU&cities[]=Bad%20Hersfeld%2C%20Hesse%2C%20DEU&cities[]=Oelde%2C%20North-Rhine-Westphalia%2C%20DEU&cities[]=Winsen%20(Luhe)%2C%20Lower-Saxony%2C%20DEU&cities[]=Rheinberg%2C%20North-Rhine-Westphalia%2C%20DEU&cities[]=Graben%2C%20Bavaria%2C%20DEU&cities[]=Hamburg%2C%20Hamburg%2C%20DEU&cities[]=M%C3%B6nchengladbach%2C%20North-Rhine-Westphalia%2C%20DEU&cities[]=Wolfhagen%2C%20Hesse%2C%20DEU&cities[]=Bremen%2C%20Bremen%2C%20DEU&cities[]=Kobern-Gondorf%2C%20Rhineland-Palatinate%2C%20DEU&cities[]=Wunstorf%2C%20Lower-Saxony%2C%20DEU&cities[]=Werne%2C%20North-Rhine-Westphalia%2C%20DEU&cities[]=Krefeld%2C%20North-Rhine-Westphalia%2C%20DEU&cities[]=Sch%C3%B6nefeld%2C%20Brandenburg%2C%20DEU&cities[]=Aachen%2C%20North-Rhine-Westphalia%2C%20DEU&cities[]=Cottbus%2C%20Brandenburg%2C%20DEU&cities[]=Dusseldorf%2C%20North-Rhine-Westphalia%2C%20DEU&cities[]=Emsb%C3%BCren%2C%20Lower-Saxony%2C%20DEU&cities[]=Erfurt%2C%20Thuringia%2C%20DEU&cities[]=Gernsheim%2C%20Hesse%2C%20DEU&cities[]=Bielefeld%2C%20North-Rhine-Westphalia%2C%20DEU&cities[]=Cloppenburg%2C%20Lower-Saxony%2C%20DEU&cities[]=Eggolsheim%2C%20Bavaria%2C%20DEU&cities[]=Freiburg%2C%20Baden-Wurttemberg%2C%20DEU&cities[]=Neuwied%2C%20Rhineland-Palatinate%2C%20DEU&cities[]=Paderborn%2C%20North-Rhine-Westphalia%2C%20DEU&cities[]=T%C3%BCbingen%2C%20Baden-Wurttemberg%2C%20DEU&cities[]=Brieselang%2C%20Brandenburg%2C%20DEU&cities[]=Garbsen%2C%20Lower-Saxony%2C%20DEU&cities[]=Hanover%2C%20Lower-Saxony%2C%20DEU&cities[]=Raunheim%2C%20Hesse%2C%20DEU&cities[]=V%C3%B6lklingen%2C%20Saarland%2C%20DEU&cities[]=Witten%2C%20North-Rhine-Westphalia%2C%20DEU&cities[]=Emden%2C%20Lower-Saxony%2C%20DEU&cities[]=Mannheim%2C%20Baden-Wurttemberg%2C%20DEU&cities[]=Nuremberg%2C%20Bavaria%2C%20DEU&cities[]=Bad%20Oldesloe%2C%20Schleswig-Holstein%2C%20DEU&cities[]=Bochum%2C%20North-Rhine-Westphalia%2C%20DEU&cities[]=Borgstedt%2C%20Schleswig-Holstein%2C%20DEU&cities[]=Essen%2C%20North-Rhine-Westphalia%2C%20DEU&cities[]=Magdeburg%2C%20Sachsen-Anhalt%2C%20DEU&cities[]=Me%C3%9Fkirch%2C%20Baden-Wurttemberg%2C%20DEU&cities[]=Schortens%2C%20Lower-Saxony%2C%20DEU&cities[]=Trierweiler%2C%20Rhineland-Palatinate%2C%20DEU&cities[]=Bayreuth%2C%20Bavaria%2C%20DEU&cities[]=Duisburg%2C%20North-Rhine-Westphalia%2C%20DEU&cities[]=Echzell%2C%20Hesse%2C%20DEU&cities[]=Kn%C3%BCllwald%2C%20Hesse%2C%20DEU&cities[]=Lengede%2C%20Lower-Saxony%2C%20DEU&cities[]=Rostock%2C%20Mecklenburg-Vorpommern%2C%20DEU&cities[]=Unna%2C%20North-Rhine-Westphalia%2C%20DEU&cities[]=Frankenberg%20(Eder)%2C%20Hesse%2C%20DEU&cities[]=Grolsheim%2C%20Rhineland-Palatinate%2C%20DEU&cities[]=N%C3%BCtzen%2C%20Schleswig-Holstein%2C%20DEU&cities[]=Pommersfelden%2C%20Bavaria%2C%20DEU&cities[]=Sonneberg%2C%20Thuringia%2C%20DEU&cities[]=Waldorf%2C%20Rhineland-Palatinate%2C%20DEU&cities[]=Wuppertal%2C%20North-Rhine-Westphalia%2C%20DEU&cities[]=Eschweiler%2C%20North-Rhine-Westphalia%2C%20DEU&cities[]=Gersthofen%2C%20Bavaria%2C%20DEU&cities[]=Hoppegarten%2C%20Brandenburg%2C%20DEU&cities[]=Kempen%2C%20North-Rhine-Westphalia%2C%20DEU&cities[]=Stuttgart%2C%20Baden-Wurttemberg%2C%20DEU&cities[]=Ulm%2C%20Baden-Wurttemberg%2C%20DEU&distanceType=Mi&radius=24km&latitude=&longitude=&loc_group_id=&loc_query=&base_query=&city=&country=&region=&county=&query_options=&'
            driver1 = webdriver.Chrome(path)
            driver1.get(url)
            
            print('page: ', page)

            obj1 = driver1.find_element_by_tag_name('body')
            obj2 = obj1.find_element_by_class_name("search-page")
            offers = obj2.find_elements_by_class_name('job-tile')

            for ad in offers:
                title = ad.find_element_by_class_name('job-title').text
                
                location_jobboard = ad.find_element_by_class_name('location-and-id').text
                location_jobboard = location_jobboard[:location_jobboard.find('|')]

                posting_date = ad.find_element_by_class_name('posting-date').text.replace('Geschaltet ', '')
                
                link = ad.find_element_by_tag_name('a').get_attribute('href')

                update = ad.find_elements_by_tag_name('p')[1].text.replace('(', '').replace(')', '').replace('aktualisiert', '').replace('Vor', '').replace('about', '').strip()
        
                if re.search('hours', update):
                    update = datetime.now().date()
                elif re.search('days', update):
                    days = int(update.replace('days', '').strip())
                    update = datetime.now().date() - timedelta(days=days)
                elif re.search('months', update):
                    months = int(update.replace('months', '').strip())
                    update = datetime.now().date() - timedelta(days=months*30)
                elif re.search('month', update):
                    update = datetime.now().date() - timedelta(days=30)
                else:
                    None

                # BS4 inside an offer
                session = requests.Session()
                headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/80.0.3987.132 Safari/537.36'}
                response_obj = session.get(link, headers=headers)
                soup_obj = BeautifulSoup(response_obj.content, 'html.parser')

                body = soup_obj.find('body')
                details = body.find('div', id = 'job-detail')
                desc = details.find('div', id = 'job-detail-body')
                desc1 = desc.find('div', class_= 'content')
                
                company_name = details.find('div', class_= 'details-line').find('p').text
                
                firstDelPos = company_name.find("J") # get the position of J - the first letter to remove
                secondDelPos = company_name.find("A") # get the position of A - the first letter to leave
                company_name = company_name.replace(company_name[firstDelPos:secondDelPos], '')
                
                firstDelPos2 = company_name.find("S") # get the position of J - the first letter to remove
                secondDelPos2 = company_name.find("|") # get the position of A - the first letter to leave
                company_name = company_name.replace(company_name[firstDelPos2:secondDelPos2 + 1], '').strip()
                
                try:    
                    sector_jobboard = desc.find_all('div', class_= 'association-content')[1].find('a').text     
                except:
                    sector_jobboard = None

                try:
                    salary = desc.find('b',  text = re.compile('€')).text
                except:
                    try:
                        salary = desc.find('br',  text = re.compile('€')).text
                    except:
                        try:
                            salary = desc.find('span',  text = re.compile('€')).text
                        except:
                            salary = None

                agency_or_direct = None
                try:
                    employment_type = desc.find('br',  text = re.compile('Arbeitsvertrag'))
                except:
                    employment_type = None

                description = desc1.find('div', class_= 'section description').find('p').text       
                require = desc1.find_all('div', class_= 'section')
                basic = require[0].text
                prefered = require[1].text

                full_description = description + '; ' + basic + '; ' + prefered
                full_description = full_description.replace('"', '').replace("'", '').replace('’', '').replace('`', '').strip()[:4000]

                company = 'Amazon'
                skills = None 
                
                # te dwie linijki z # służyły do różnych testów. Nie usuwałem, bo nie pamiętam dokładnie, czy można ich jeszcze użyć
                # new = self.printing(title, company_name, location_jobboard, posting_date, sector_jobboard, salary, agency_or_direct, employment_type, full_description, skills, link, update)
                # self.printing to metoda stworza w jobboardzie przeze mnie do drukowania. Możesz dać tak, żeby było normalnie do bazy rzucane.
                new = self.printing(title, company_name, location_jobboard, posting_date, sector_jobboard, salary, agency_or_direct, employment_type, full_description, skills, link)
                # new_list = ((new[0], company, new[1], new[2], new[3], new[4], new[5], new[6], new[7], new[8], new[9], new[10], new[11], new[12],))

                num = nums
                print(num)
                ws['A' + str(num)] = new[0]
                ws['B' + str(num)] = company
                ws['C' + str(num)] = new[1]
                ws['D' + str(num)] = new[2]
                ws['E' + str(num)] = new[3]
                ws['F' + str(num)] = new[4]
                ws['G' + str(num)] = new[5]
                ws['H' + str(num)] = new[6]
                ws['I' + str(num)] = new[7]
                ws['J' + str(num)] = new[8]
                ws['K' + str(num)] = new[9]
                ws['L' + str(num)] = new[10]
                ws['M' + str(num)] = new[11]
                ws['N' + str(num)] = update
                nums += 1
                wb.save("Amazon_NEW.xlsx")
        driver1.close()                                
    # df = pd.DataFrame(total, columns = ['Scraping Date', 'Company', 'Title', 'Company Name', 'Location', 'Date', 'Sector', 'Salary', 'Agency or Direct', 'Employment Type', 'Description', 'Skills', 'Link']) 
    # df.to_excel('Amazon_NEW.xlsx')
        wb.save("Amazon_NEW2.xlsx")

if __name__ == "__main__":
    amazon_de = AmazonScraper('Germany', 'amazon.de', 'EUR')
    amazon_de.explore_ads()
