# this one is useful
from flashtext import KeywordProcessor
from unidecode import unidecode
from openpyxl import load_workbook

wb = load_workbook(r'path\SelectedCompanies - Top 50.xlsx')

sheet_obj = wb.active 
max_col = sheet_obj.max_column 
max_row = sheet_obj.max_row
num = 1
sheet_obj['N1'] = 'Skills'
for i in range(1, max_row + 1): 
    full_description = sheet_obj.cell(row = i, column = 13).value
    skills = ''

    skills_path = r'path\skills.txt'

    with open(skills_path, 'r', encoding='utf8') as f:
        skills_list = f.read()
        skills_list = skills_list.split('\n')
        skills_keyword_processor_short = KeywordProcessor(case_sensitive=True)
        skills_keyword_processor_long = KeywordProcessor(case_sensitive=False)
        for skill in skills_list:
            if len(skill.strip()) > 4:
                skills_keyword_processor_long.add_keyword(skill.strip())
            else:
                skills_keyword_processor_short.add_keyword(skill.strip())

    if skills:
        skills = skills.replace('"', '').replace("'", '').replace('’', '').replace('`', '').strip()[:4000]
    else:
        skills = ''
        try:
            skills = ';'.join(list(set(skills_keyword_processor_short.extract_keywords(unidecode(full_description.replace('i.d.R.', ''))))) + list(set(skills_keyword_processor_long.extract_keywords(full_description))))
        except:
            None
        skills = skills[:4000]
        if len(skills) == 0:
            skills = None
    num += 1
    sheet_obj['N' + str(i)] = skills
    print(num)
wb.save("Excel_skills.xlsx")
